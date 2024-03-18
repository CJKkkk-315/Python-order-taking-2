import os
import time
import json
import random
import requests
from fake_useragent import UserAgent
from openpyxl import Workbook, load_workbook
data = []
cities = ['广东', '北京', '上海', '天津', '重庆', '云南', '黑龙江', '内蒙古', '吉林',
          '宁夏', '安徽', '山东', '山西', '四川', '广西', '新疆', '江苏', '江西',
          '河北', '河南', '浙江', '海南', '湖北', '湖南', '澳门', '甘肃',
          '福建', '西藏', '贵州', '辽宁', '陕西', '青海', '香港', '台湾']
def get_city_scenic(city, page):
    ua = UserAgent(verify_ssl=False)
    headers = {
        'authority': 'piao.qunar.com',
        'cache-control': 'max-age=0',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Opera";v="85"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36 OPR/85.0.4341.60',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'cookie': 'QN1=00008d802eb4407c7850dd92; QN300=organic; QN99=2070; qunar-assist={%22version%22:%2220211215173359.925%22%2C%22show%22:false%2C%22audio%22:false%2C%22speed%22:%22middle%22%2C%22zomm%22:1%2C%22cursor%22:false%2C%22pointer%22:false%2C%22bigtext%22:false%2C%22overead%22:false%2C%22readscreen%22:false%2C%22theme%22:%22default%22}; QN205=organic; QN277=organic; QN267=0623206714e3b569d6; csrfToken=weCkd8DWqzc29Wyl7YoFRIjB7TC78IgF; QN269=EA328D21BC8011EC895AFA163E763FD9; _i=VInJOmqoD6pUOypqYqhLV2riSHYq; _vi=U6221NK6MZ5CFz86Ee0L7zEK1JGTn3QNir5FDimAOwLZvUTph2VPu4EU9h5C0GcV5AIVvNsY6I1zDNu8JlMeTuLUjfKxMfjTID2L6uaWpsO20A2mRpfKKRWcOC5ZsHu-3ywKWZ0OqGJRuiT44rn6jIHR0cwNZxsnvNe1AWu4M6-Q; QN601=9a98101fa4db62811b0762c2d7aac1e6; QN48=00008d802f10407c7a086b0a; quinn=4860de8fc62b970a4a21d3097b76f217af609d4f50e4123cecda580239f39402b5f0ddb803a0ff709cfce4f501055247; fid=642f58a1-2e78-4741-abc6-fdedfda84240; QN271=33092d82-4fd3-4e84-8814-7d38183d99cc; RT=; JSESSIONID=40F684FE943F90F7B889665F527F9F17; JSESSIONID=42D75318BEC4EC99FD0079E829AE4C4E'
    }
    url = f'https://piao.qunar.com/ticket/list.json?keyword={city}&region=&from=mpl_search_suggest&sort=pp&page={page}'
    result = requests.get(url, headers=headers, timeout=10)
    result.raise_for_status()
    return  get_scenic_info(city, result.text)
def get_scenic_info(city, response):
    response_info = json.loads(response)
    sight_list = response_info['data']['sightList']
    one_city_scenic = []
    for sight in sight_list:
        scenic = []
        name = sight['sightName'] # 景点名称
        star = sight.get('star', None) # 星级
        score = sight.get('score', 0) # 评分
        price = sight.get('qunarPrice', 0) # 价格
        sale = sight.get('saleCount', 0) # 销量
        districts = sight.get('districts', None) # 省，市，区
        point = sight.get('point', None) # 坐标
        intro = sight.get('intro', None) # 简介
        free = sight.get('free', True) # 是否免费
        address = sight.get('address', None) # 具体地址
        scenic.append(city)
        scenic.append(name)
        scenic.append(star)
        scenic.append(score)
        scenic.append(price)
        scenic.append(sale)
        scenic.append(districts)
        scenic.append(point)
        scenic.append(intro)
        scenic.append(free)
        scenic.append(address)
        one_city_scenic.append(scenic)
    return one_city_scenic
def get_city_info(cities, pages):
    for city in cities:
        one_city_info = []
        for page in range(1, pages+1):
            try:
                print(f'正在爬取-{city}(省/市), 第{page}页景点数据...')
                time.sleep(random.uniform(0.8,1.5))
                one_page_info = get_city_scenic(city, page)
            except:
                continue
            if one_page_info:
                one_city_info += one_page_info
        insert2excel(city+'.xlsx',one_city_info)
def insert2excel(filepath,allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = ['城市','名称','星级','评分','价格','销量','省/市/区','坐标','简介','是否免费','具体地址']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(3)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        for info in allinfo:
            ws.append(info)
        wb.save(filepath)
        return True
    except:
        return False
get_city_info(cities,5)