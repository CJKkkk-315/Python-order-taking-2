from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('11.xlsx')

sh = wb.worksheets
data = []
table = sh[0]
for i in table.rows:
    aw = []
    for j in i:
        aw.append(j.value)
    data.append(aw[:8])
sum = 0
ans = []
for i in data[1:]:
    gm = round(i[0]*i[1],5)
    zhi = round(sum + gm,5)
    sum += gm
    if zhi < i[4]/5:
        shuchu = round(zhi/150,5)
    elif zhi >= i[4]/5 and zhi < i[4]/3:
        shuchu = round((zhi-1)/150,5)
        sum -= i[5]
    elif zhi >= i[4]/3:
        shuchu = round((zhi-4)/150,5)
        sum -= i[5] + i[6]
    ans.append([i[0],i[1],gm,zhi,150,1,3,shuchu])
head = ['高度','面积','高度*面积','值','容积','最小','最大','输出']
res = [head] + ans
wb = Workbook()
wb.create_sheet(index=1, title="sheet2")
ws = wb.active
for i in range(len(res)):
    for j in range(len(res[i])):
        ws.cell(i+1, j+1).value = res[i][j]
wb.save('res.xlsx')

