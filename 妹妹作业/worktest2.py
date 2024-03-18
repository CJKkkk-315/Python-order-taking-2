import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook(r"名单.xlsx")
ws = wb['Sheet1']
wb_out = Workbook()
ws_out = wb_out.active
num = []
name = []
studentID = []
for each_row in ws.iter_rows(min_row=0, min_col=0, max_row=31, max_col=3):
    num.append(each_row[0].value)
    name.append(each_row[1].value)
    studentID.append(each_row[2].value)
for i in range(len(num)):
    if name[i] == '徐东海':
        num.insert(1, num[i])
        name.insert(1, name[i])
        studentID.insert(1, studentID[i])
        del num[i+1]
        del name[i+1]
        del studentID[i+1]
for  x  in  range(1,len(num)+1):
        ws_out.cell(x,  1,  value  =  num[x-1])
        ws_out.cell(x,  2,  value  =  name[x-1])
        ws_out.cell(x,  3,  value  =  studentID[x-1])
wb_out.save(filename  =  "新名单.csv")
